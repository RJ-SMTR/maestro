from repositories.helpers.io import connect_ftp, get_blob
from dagster import (
    solid,
    Output,
    OutputDefinition,
    composite_solid,
)

import io
import requests
import json
import pendulum
import pandas as pd
from pathlib import Path
import shutil
import os
from openpyxl import load_workbook
import re
from google.cloud import storage

import basedosdados as bd
from basedosdados import Table

# Temporario, essa funcao vai ser incorporada a base dos dados
from repositories.helpers.storage import StoragePlus
from repositories.helpers.io import get_credentials_from_env


@solid(
    output_defs=[
        OutputDefinition(name="filename"),
        OutputDefinition(name="partitions"),
    ],
    required_resource_keys={"timezone_config"},
)
def create_current_datetime_partition(context):
    timezone = context.resources.timezone_config["timezone"]

    capture_time = pendulum.now(timezone)
    date = capture_time.strftime("%Y-%m-%d")
    hour = capture_time.strftime("%H")
    filename = capture_time.strftime("%Y-%m-%d-%H-%M-%S")

    partitions = f"data={date}/hora={hour}"

    yield Output(filename, output_name="filename")
    yield Output(partitions, output_name="partitions")


@solid(
    required_resource_keys={"basedosdados_config"},
)
def get_file_path_and_partitions(context, filename, partitions, table_id=None):

    # If not specific table_id, use resource one
    if not table_id:
        table_id = context.resources.basedosdados_config["table_id"]
    dataset_id = context.resources.basedosdados_config["dataset_id"]

    # Get data folder from environment variable
    data_folder = os.getenv("DATA_FOLDER", "data")

    file_path = f"{os.getcwd()}/{data_folder}/{{mode}}/{dataset_id}/{table_id}/{partitions}/{filename}.{{filetype}}"
    context.log.info(f"creating file path {file_path}")

    return file_path


@solid(
    output_defs=[
        OutputDefinition(name="filename"),
        OutputDefinition(name="filetype"),
        OutputDefinition(name="file_path"),
        OutputDefinition(name="partitions"),
    ],
)
def parse_file_path_and_partitions(context, bucket_path):

    # Parse bucket to get mode, dataset_id, table_id and filename
    path_list = bucket_path.split("/")
    dataset_id = path_list[1]
    table_id = path_list[2]
    filename = path_list[-1].split(".")[0]
    filetype = path_list[-1].split(".")[1]

    # Parse bucket to get partitions
    partitions = re.findall("\/([^\/]*?)=(.*?)(?=\/)", bucket_path)
    partitions = "/".join(["=".join([field for field in item])
                          for item in partitions])

    # Get data folder from environment variable
    data_folder = os.getenv("DATA_FOLDER", "data")

    folder = f"{os.getcwd()}/{data_folder}/{{mode}}/{dataset_id}/{table_id}/"
    file_path = f"{folder}/{partitions}/{filename}.{{filetype}}"
    context.log.info(f"creating file path {file_path}")

    yield Output(filename, output_name="filename")
    yield Output(filetype, output_name="filetype")
    yield Output(file_path, output_name="file_path")
    yield Output(partitions, output_name="partitions")


@solid(required_resource_keys={'basedosdados_config', 'timezone_config'})
def upload_logs_to_bq(context, timestamp, error):

    dataset_id = context.resources.basedosdados_config['dataset_id']
    table_id = context.resources.basedosdados_config['table_id'] + "_logs"

    filepath = Path(
        f"{timestamp}/{table_id}/data={pendulum.parse(timestamp).date()}/{table_id}_{timestamp}.csv")
    # create partition directory
    filepath.parent.mkdir(exist_ok=True, parents=True)
    # create dataframe to be uploaded
    df = pd.DataFrame(
        {"timestamp_captura": [pd.to_datetime(timestamp)], "sucesso": [
            error is None], "erro": [error]}
    )
    # save local
    df.to_csv(filepath, index=False)
    # BD Table object
    tb = Table(table_id, dataset_id)
    # create and publish if table does not exist, append to it otherwise
    if not tb.table_exists("staging"):
        tb.create(
            path=f"{timestamp}/{table_id}",
            if_table_exists="replace",
            if_storage_data_exists="replace",
            if_table_config_exists="pass",
        )
    elif not tb.table_exists("prod"):
        tb.publish(if_exists="replace")
    else:
        tb.append(filepath=f"{timestamp}/{table_id}", if_exists='replace')

    # delete local file
    shutil.rmtree(f"{timestamp}")


@solid(
    output_defs=[
        OutputDefinition(name="data", is_required=False),
        OutputDefinition(name="timestamp", is_required=False),
        OutputDefinition(name="error", is_required=False)],
    required_resource_keys={"basedosdados_config", "timezone_config"},
)
def get_raw(context, url, headers=None, kind=None):

    data = None
    error = None
    timestamp = pendulum.now(context.resources.timezone_config["timezone"])

    if kind == "sppo":
        url = url + "?" +"".join([k + "=" + v for k, v in headers.items()])
        headers = None

    try:
        data = requests.get(url, headers=headers, timeout=60)
    except requests.exceptions.ReadTimeout as e:
        error = e
    except Exception as e:
        error = f"Unknown exception while trying to fetch data from {url}: {e}"

    if data is None:
        if error is None:
            error = "Data from API is none!"

    if error:
        yield Output(timestamp.isoformat(), output_name="timestamp")
        yield Output(error, output_name="error")
    elif data.ok:
        yield Output(data, output_name="data")
        yield Output(timestamp.isoformat(), output_name="timestamp")
        yield Output(error, output_name="error")
    else:
        error = f"Requests failed with error {data.status_code}"
        yield Output(timestamp.isoformat(), output_name="timestamp")
        yield Output(error, output_name="error")


@solid
def save_raw_local(context, data, file_path, mode="raw"):

    _file_path = file_path.format(mode=mode, filetype="json")
    Path(_file_path).parent.mkdir(parents=True, exist_ok=True)
    json.dump(data.json(), Path(_file_path).open("w"))

    return _file_path


@solid
def save_treated_local(context, df, file_path, mode="staging"):

    _file_path = file_path.format(mode=mode, filetype="csv")
    Path(_file_path).parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(_file_path, index=False)

    return _file_path


@solid(
    required_resource_keys={"basedosdados_config"},
)
def get_file_from_storage(
    context,
    file_path,
    filename,
    partitions,
    mode="raw",
    filetype="xlsx",
    uploaded=True,
    table_id=None,
):

    # Download from storage
    # If not specific table_id, use resource one
    if not table_id:
        table_id = context.resources.basedosdados_config["table_id"]
    dataset_id = context.resources.basedosdados_config["dataset_id"]

    _file_path = file_path.format(mode=mode, filetype=filetype)

    st = StoragePlus(table_id=table_id, dataset_id=dataset_id)
    context.log.debug(f"File path: {_file_path}")
    context.log.debug(f"filename: {filename}")
    context.log.debug(f"partition: {partitions}")
    context.log.debug(f"mode: {mode}")
    st.download(
        filename=filename + "." + filetype,
        file_path=_file_path,
        partitions=partitions,
        mode=mode,
        if_exists="replace",
    )
    return _file_path


def save_local_as_bd(
    data,
    data_folder,
    file_name,
    dataset_id,
    table_id,
    mode,
    filetype,
    partitions=None,
):

    file_name = f"{file_name}.{filetype}"

    if partitions == None:
        file_path = f"{data_folder}/{mode}/{dataset_id}/{table_id}/"
    else:
        file_path = f"{data_folder}/{mode}/{dataset_id}/{table_id}/{partitions}/"

    return save_local(data, file_path, file_name)


def save_local(data, file_path="tmp", file_name="tmp"):
    """Saves data locally."""

    if file_path == "tmp":
        file_path = "TMP/"

    file_path = Path(file_path) / Path(file_name)

    Path(file_path).parent.mkdir(parents=True, exist_ok=True)

    if isinstance(data, pd.DataFrame):
        data.to_csv(file_path, index=False)
    elif isinstance(data, dict):
        json.dump(data, Path(file_path).open("w"))
    else:
        Path(file_path).open("w").write(data)

    return file_path


def delete_file(file_path):
    return Path(file_path).unlink(missing_ok=True)


@solid
def download_file_from_ftp(ftp_path: str, local_path: str) -> None:
    """Downloads a file from FTP to the local storage"""
    Path(local_path).parent.mkdir(parents=True, exist_ok=True)
    ftp_client = connect_ftp(os.getenv("FTPS_HOST"), os.getenv(
        "FTPS_USERNAME"), os.getenv("FTPS_PWD"))
    ftp_client.retrbinary(
        "RETR " + ftp_path, open(local_path, "wb").write)
    ftp_client.quit()


@solid(
    required_resource_keys={"basedosdados_config"},
)
def upload_file_to_storage(
    context, file_path, partitions=None, mode="raw", table_id=None
):

    # Upload to storage
    # If not specific table_id, use resource one
    if not table_id:
        table_id = context.resources.basedosdados_config["table_id"]
    dataset_id = context.resources.basedosdados_config["dataset_id"]

    st = bd.Storage(table_id=table_id, dataset_id=dataset_id)

    context.log.debug(
        f"Uploading file {file_path} to mode {mode} with partitions {partitions}"
    )
    st.upload(path=file_path, mode=mode,
              partitions=partitions, if_exists="replace")

    return True


@solid(
    required_resource_keys={"basedosdados_config"},
)
def upload_blob_to_storage(
    context, blob_path, partitions=None, mode="raw", table_id=None, bucket_name="", credential_mode="staging"
):
    # Extracted from basedosdados
    def _resolve_partitions(partitions):
        if isinstance(partitions, dict):
            return "/".join(f"{k}={v}" for k, v in partitions.items()) + "/"
        elif isinstance(partitions, str):
            if partitions.endswith("/"):
                partitions = partitions[:-1]
            # If there is no partition
            if len(partitions) == 0:
                return ""
            # It should fail if there is folder which is not a partition
            try:
                # check if it fits rule
                {b.split("=")[0]: b.split("=")[1]
                 for b in partitions.split("/")}
            except IndexError:
                raise Exception(
                    f"The path {partitions} is not a valid partition")
            return partitions + "/"
        else:
            raise Exception(
                f"Partitions format or type not accepted: {partitions}")

    if not table_id:
        table_id = context.resources.basedosdados_config["table_id"]
    dataset_id = context.resources.basedosdados_config["dataset_id"]
    credentials = get_credentials_from_env(mode=credential_mode)
    client = storage.Client(credentials=credentials)
    blob_name = f"{mode}/{dataset_id}/{table_id}/"
    if partitions is not None:
        blob_name += _resolve_partitions(partitions=partitions)
    blob_name += blob_path.split("/")[-1]
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(blob_name)
    input_blob = get_blob(blob_path, bucket_name, mode=credential_mode)
    context.log.debug(
        f"Uploading blob {blob_path} to mode {mode} with partitions {partitions}"
    )
    blob.upload_from_file(io.BytesIO(input_blob.download_as_bytes()))


@solid
def delete_xls_header(context, file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    # Delete current header
    ws.delete_rows(1)
    wb.save(file_path)
    return file_path


@solid
def set_xls_header(context, file_path, header):
    wb = load_workbook(file_path)
    ws = wb.active
    ws.insert_rows(1)
    for idx, column in enumerate(header):
        ws.cell(row=1, column=idx + 1).value = column
    wb.save(file_path)

    return file_path


@composite_solid
def set_header(file_path, header):
    file_path = delete_xls_header(file_path)
    file_path = set_xls_header(file_path, header)
    return file_path
